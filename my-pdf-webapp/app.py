import os
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
import pdfplumber

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
OUTPUT_FILE = 'Final_CMR.xlsx'
TEMPLATE_FILE = 'template.xlsx'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def extract_data_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text()

    lines = text.split('\n')

    load_no = ''
    customer_name = ''
    destination_address = ''
    destination_location = ''
    total_qty = ''
    total_vol = ''
    total_gw = ''
    total_cases = ''

    for idx, line in enumerate(lines):
        if 'Load No:' in line:
            load_no = line.split(':')[1].strip()
        if 'Customer Name' in line:
            customer_name = line.split('Customer Name')[1].strip()
        if 'Final Destination Address' in line:
            destination_address = '\n'.join(lines[idx+1:idx+4])
        if 'Destination Location' in line:
            destination_location = line.split('Destination Location')[1].strip()
        if 'TOTAL' in line:
            parts = line.split()
            if len(parts) >= 5:
                total_qty = parts[1]
                total_vol = parts[2]
                total_gw = parts[3]
                total_cases = parts[4]

    return {
        'load_no': load_no,
        'customer_name': customer_name,
        'destination_address': destination_address,
        'destination_location': destination_location,
        'totals': {
            'qty': total_qty,
            'vol': total_vol,
            'gw': total_gw,
            'cases': total_cases
        }
    }

def fill_excel(data):
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    ws['C20'] = data['load_no']
    ws['C7'] = data['customer_name']

    address_lines = data['destination_address'].split('\n')
    ws['C8'] = address_lines[0] if len(address_lines) > 0 else ''
    ws['C9'] = address_lines[1] if len(address_lines) > 1 else ''
    ws['C10'] = address_lines[2] if len(address_lines) > 2 else ''

    ws['C13'] = data['destination_location']

    ws['I22'] = data['totals']['qty']
    ws['J22'] = data['totals']['vol']
    ws['K22'] = data['totals']['gw']
    ws['L22'] = data['totals']['cases']

    wb.save(OUTPUT_FILE)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'pdf_file' not in request.files:
            return "No file part"
        file = request.files['pdf_file']
        if file.filename == '':
            return "No selected file"
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)

        data = extract_data_from_pdf(file_path)
        fill_excel(data)

        return send_file(OUTPUT_FILE, as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
